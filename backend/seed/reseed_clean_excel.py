import openpyxl
import asyncio
import asyncpg
import uuid
import os

filepath = "/app/seed/devotees.xlsx" if os.path.exists("/app/seed/devotees.xlsx") else "/app/devotees.xlsx"
DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://kovil_user:kovil_pass@postgres:5432/sembukutty_kovil"
).replace("postgresql+asyncpg://", "postgresql://")

wb = openpyxl.load_workbook(filepath, data_only=True)
sheet = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active

members = []
for r in range(2, sheet.max_row + 1):
    name = sheet.cell(r, 2).value
    phone = sheet.cell(r, 3).value
    address = sheet.cell(r, 4).value

    if name is not None and str(name).strip():
        name_clean = str(name).strip()
        phone_clean = str(phone).strip() if phone is not None else ""
        if phone_clean.endswith(".0"):
            phone_clean = phone_clean[:-2]
        address_clean = str(address).strip() if address is not None else ""

        members.append((name_clean, phone_clean, address_clean))

async def reseed_clean():
    conn = await asyncpg.connect(DATABASE_URL)
    # Clear members table to replace with clean UTF-8 Excel data
    await conn.execute("DELETE FROM members")
    print("Cleared existing members table.")

    inserted = 0
    for name, phone, address in members:
        await conn.execute(
            "INSERT INTO members (id, name, phone, address, created_at, updated_at) "
            "VALUES ($1, $2, $3, $4, NOW(), NOW())",
            str(uuid.uuid4()), name, phone, address
        )
        inserted += 1

    total = await conn.fetchval("SELECT COUNT(*) FROM members")
    await conn.close()

    print(f"Reseeded {inserted} members cleanly from Excel. Total in DB: {total}")

if __name__ == "__main__":
    asyncio.run(reseed_clean())
