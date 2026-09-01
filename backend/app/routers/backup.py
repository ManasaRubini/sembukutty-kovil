import io
from typing import Optional
from datetime import datetime, timezone, date as DateType
from fastapi import APIRouter, Depends, HTTPException, Response, Query
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete
from app.database import get_db
from app.models.staff import Staff
from app.models.member import Member
from app.models.opening_balance import OpeningBalance
from app.models.transaction import Transaction
from app.models.document_sequence import DocumentSequence
from app.routers.auth import require_admin
import openpyxl

router = APIRouter(prefix="/api/backup", tags=["backup"])


from app.schemas import ResetAccountingRequest
import uuid


@router.post("/reset-accounting-data")
async def reset_accounting_data(
    body: Optional[ResetAccountingRequest] = None,
    db: AsyncSession = Depends(get_db),
    admin: dict = Depends(require_admin),
):
    """
    Admin exclusive reset: Deletes all receipts, vouchers, transfer notes,
    resets serial document counters (R-00000, V-00000, T-00000), and sets new opening balances.
    IMPORTANT: Preserves the entire Devotees (members) list and Staff accounts intact!
    """
    # 1. Delete all billing transactions
    await db.execute(delete(Transaction))

    # 2. Reset document sequences
    await db.execute(delete(DocumentSequence))

    # 3. Clear existing opening balances
    await db.execute(delete(OpeningBalance))

    # 4. Set new initial opening balances specified by Admin
    bank_bal = body.bank_balance if body else 0.0
    cash_bal = body.cash_balance if body else 0.0

    new_ob = OpeningBalance(
        id=str(uuid.uuid4()),
        bank_balance=bank_bal,
        cash_balance=cash_bal,
    )
    db.add(new_ob)

    await db.commit()

    return {
        "message": f"All receipts, vouchers, and transaction data reset successfully. Opening Bank: ₹{bank_bal:,.2f}, Opening Cash: ₹{cash_bal:,.2f}. Devotees list preserved.",
        "success": True,
        "opening_bank_balance": bank_bal,
        "opening_cash_balance": cash_bal,
    }


@router.get("/export-excel")
async def export_excel_backup(
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    prefix: Optional[str] = Query("backup-1"),
    db: AsyncSession = Depends(get_db),
):
    """
    Generates and returns a complete, beautifully formatted Excel workbook (.xlsx)
    containing all temple accounting data, transactions (filtered by date range if provided),
    devotees, and balances.
    """
    staff_res = await db.execute(select(Staff))
    members_res = await db.execute(select(Member).order_by(Member.name))
    ob_res = await db.execute(select(OpeningBalance).limit(1))

    q_txns = select(Transaction)
    if date_from:
        q_txns = q_txns.where(Transaction.date >= DateType.fromisoformat(date_from))
    if date_to:
        q_txns = q_txns.where(Transaction.date <= DateType.fromisoformat(date_to))
    q_txns = q_txns.order_by(Transaction.created_at.desc())

    txns_res = await db.execute(q_txns)

    staff_list = staff_res.scalars().all()
    member_list = members_res.scalars().all()
    ob = ob_res.scalar_one_or_none()
    txn_list = txns_res.scalars().all()

    staff_map = {s.id: s.name for s in staff_list}

    wb = openpyxl.Workbook()

    # Sheet 1: Transactions
    ws_txns = wb.active
    ws_txns.title = "Transactions"
    ws_txns.append([
        "Serial No", "Date", "Transaction Type", "Billed By Staff",
        "Member / Devotee", "Phone Number", "Address",
        "Purpose / Remarks", "Paid To", "Payment Mode", "Amount (INR)"
    ])

    for t in txn_list:
        type_label = (
            "Tax Collection" if t.type == "tax"
            else "Donation Collection" if t.type == "donation"
            else "Expense" if t.type == "expense"
            else "Transfer"
        )
        staff_name = staff_map.get(t.staff_id, t.staff_id or "—")
        ws_txns.append([
            t.serial_number or "—",
            str(t.date),
            type_label,
            staff_name,
            t.member_name or "—",
            t.member_phone or "—",
            t.address or "—",
            t.purpose or t.remarks or "—",
            t.paid_to or "—",
            t.mode or "—",
            float(t.amount),
        ])

    # Sheet 2: Devotees
    ws_members = wb.create_sheet(title="Devotees")
    ws_members.append(["S.No", "Devotee Name", "Phone Number", "Address"])
    for idx, m in enumerate(member_list, 1):
        ws_members.append([idx, m.name, m.phone or "—", m.address or "—"])

    # Sheet 3: Balances
    ws_bal = wb.create_sheet(title="Opening Balances")
    ws_bal.append(["Item", "Amount (INR)"])
    if ob:
        ws_bal.append(["Opening Bank Balance", float(ob.bank_balance)])
        ws_bal.append(["Opening Cash Balance", float(ob.cash_balance)])
    else:
        ws_bal.append(["Opening Bank Balance", 0.0])
        ws_bal.append(["Opening Cash Balance", 0.0])

    # Sheet 4: Staff
    ws_staff = wb.create_sheet(title="Billing Members")
    ws_staff.append(["Staff ID", "Staff Name", "Phone", "Email", "Status"])
    for s in staff_list:
        status_str = "Active" if (s.is_active and s.is_approved) else "Inactive/Pending"
        ws_staff.append([s.id, s.name, s.phone or "—", s.email or "—", status_str])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    today_str = datetime.now().strftime("%Y-%m-%d")
    clean_prefix = prefix or "backup-1"
    if date_from and date_to:
        filename = f"{clean_prefix}-{date_from}-to-{date_to}.xlsx"
    else:
        filename = f"{clean_prefix}-{today_str}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export")
async def export_backup(db: AsyncSession = Depends(get_db)):
    staff_res = await db.execute(select(Staff))
    members_res = await db.execute(select(Member))
    ob_res = await db.execute(select(OpeningBalance).limit(1))
    txns_res = await db.execute(select(Transaction))
    seqs_res = await db.execute(select(DocumentSequence))

    staff_list = staff_res.scalars().all()
    member_list = members_res.scalars().all()
    ob = ob_res.scalar_one_or_none()
    txn_list = txns_res.scalars().all()
    seq_list = seqs_res.scalars().all()

    data = {
        "exportedAt": datetime.now(timezone.utc).isoformat(),
        "version": "1.0",
        "staff": [
            {"id": s.id, "name": s.name, "is_active": s.is_active,
             "created_at": s.created_at.isoformat(), "updated_at": s.updated_at.isoformat()}
            for s in staff_list
        ],
        "members": [
            {"id": m.id, "name": m.name, "phone": m.phone, "address": m.address,
             "created_at": m.created_at.isoformat(), "updated_at": m.updated_at.isoformat()}
            for m in member_list
        ],
        "openingBalance": {
            "id": ob.id,
            "bank_balance": float(ob.bank_balance),
            "cash_balance": float(ob.cash_balance),
            "cash_holder_staff_id": ob.cash_holder_staff_id,
            "created_at": ob.created_at.isoformat(),
        } if ob else None,
        "transactions": [
            {
                "id": t.id, "staff_id": t.staff_id, "type": t.type,
                "date": str(t.date), "amount": float(t.amount), "mode": t.mode,
                "member_id": t.member_id, "member_name": t.member_name,
                "member_phone": t.member_phone, "address": t.address,
                "purpose": t.purpose, "remarks": t.remarks,
                "paid_to": t.paid_to, "direction": t.direction,
                "serial_number": t.serial_number,
                "created_at": t.created_at.isoformat(),
            }
            for t in txn_list
        ],
        "documentSequences": [
            {"id": s.id, "document_type": s.document_type, "current_number": s.current_number}
            for s in seq_list
        ],
    }
    return JSONResponse(content=data)


@router.post("/import")
async def import_backup(payload: dict, db: AsyncSession = Depends(get_db)):
    required_keys = {"staff", "members", "transactions", "documentSequences"}
    if not required_keys.issubset(payload.keys()):
        raise HTTPException(status_code=400, detail=f"Invalid backup: missing keys {required_keys - payload.keys()}")

    if not isinstance(payload["staff"], list):
        raise HTTPException(status_code=400, detail="Invalid backup: staff must be a list")
    if not isinstance(payload["transactions"], list):
        raise HTTPException(status_code=400, detail="Invalid backup: transactions must be a list")

    await db.execute(delete(Transaction))
    await db.execute(delete(DocumentSequence))
    await db.execute(delete(OpeningBalance))

    for s in payload["staff"]:
        existing = await db.execute(select(Staff).where(Staff.id == s["id"]))
        if not existing.scalar_one_or_none():
            db.add(Staff(
                id=s["id"], name=s["name"],
                is_active=s.get("is_active", True),
            ))

    ob = payload.get("openingBalance")
    if ob:
        db.add(OpeningBalance(
            id=ob["id"],
            bank_balance=ob["bank_balance"],
            cash_balance=ob["cash_balance"],
            cash_holder_staff_id=ob.get("cash_holder_staff_id"),
        ))

    from datetime import date as DateType
    for t in payload["transactions"]:
        db.add(Transaction(
            id=t["id"], staff_id=t.get("staff_id"), type=t["type"],
            date=DateType.fromisoformat(t["date"]),
            amount=t["amount"], mode=t.get("mode"),
            member_id=t.get("member_id"), member_name=t.get("member_name", ""),
            member_phone=t.get("member_phone", ""), address=t.get("address", ""),
            purpose=t.get("purpose", ""), remarks=t.get("remarks", ""),
            paid_to=t.get("paid_to", ""), direction=t.get("direction"),
            serial_number=t.get("serial_number"),
        ))

    for s in payload.get("documentSequences", []):
        db.add(DocumentSequence(
            document_type=s["document_type"],
            current_number=s["current_number"],
        ))

    await db.commit()
    return {"message": "Backup imported successfully"}
