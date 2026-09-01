"""
Import devotees / members from Excel file into PostgreSQL database.
Usage: python -m seed.import_excel [path_to_excel]
"""
import asyncio
import os
import sys
import uuid
import openpyxl
import asyncpg

DEFAULT_EXCEL_PATH = os.environ.get(
    "EXCEL_PATH",
    "/downloads/SEMBUKUTTY KOVIL ADDRESS MARCH 2026.xlsx"
)
if not os.path.exists(DEFAULT_EXCEL_PATH):
    # Try local windows path if running locally
    DEFAULT_EXCEL_PATH = r"C:\Users\manas\Downloads\SEMBUKUTTY KOVIL ADDRESS MARCH 2026.xlsx"

DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://kovil_user:kovil_pass@postgres:5432/sembukutty_kovil"
).replace("postgresql+asyncpg://", "postgresql://")


def parse_excel_members(filepath: str) -> list[tuple[str, str, str]]:
    if not os.path.exists(filepath):
        print(f"Excel file not found at: {filepath}")
        return []

    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active

    members = []
    for r in range(2, sheet.max_row + 1):
        name = sheet.cell(r, 2).value
        phone = sheet.cell(r, 3).value
        address = sheet.cell(r, 4).value

        if name is not None and str(name).strip():
            name_clean = str(name).strip()
            phone_clean = str(phone).strip() if phone is not None else ""
            if phone_clean.endswith(".0"):
                phone_clean = phone_clean[:-2]
            address_clean = str(address).strip() if address is not None else ""

            members.append((name_clean, phone_clean, address_clean))

    return members


async def main():
    excel_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_EXCEL_PATH
    print(f"Parsing Excel file: {excel_path}")
    members = parse_excel_members(excel_path)
    if not members:
        print("No member records found to import.")
        return

    print(f"Loaded {len(members)} devotee records from Excel.")
    print(f"Connecting to database: {DATABASE_URL}")
    conn = await asyncpg.connect(DATABASE_URL)

    inserted = 0
    updated = 0

    for name, phone, address in members:
        existing = await conn.fetchrow("SELECT id, phone, address FROM members WHERE name = $1", name)
        if existing:
            if existing['phone'] != phone or existing['address'] != address:
                await conn.execute(
                    "UPDATE members SET phone = $1, address = $2, updated_at = NOW() WHERE id = $3",
                    phone, address, existing['id']
                )
                updated += 1
        else:
            await conn.execute(
                "INSERT INTO members (id, name, phone, address, created_at, updated_at) "
                "VALUES ($1, $2, $3, $4, NOW(), NOW())",
                str(uuid.uuid4()), name, phone, address
            )
            inserted += 1

    total_in_db = await conn.fetchval("SELECT COUNT(*) FROM members")
    await conn.close()

    print("\n--- IMPORT SUMMARY ---")
    print(f"New devotees inserted:      {inserted}")
    print(f"Existing devotees updated:  {updated}")
    print(f"Total devotees in DB:       {total_in_db}")
    print("-----------------------\n")


if __name__ == "__main__":
    asyncio.run(main())
