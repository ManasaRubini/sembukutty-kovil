import io
import uuid
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, or_
from app.database import get_db
from app.models.member import Member
from app.schemas import MemberCreate, MemberUpdate, MemberOut
from app.routers.auth import get_current_user
import openpyxl

router = APIRouter(prefix="/api/members", tags=["members"])


@router.get("", response_model=list[MemberOut])
async def list_members(
    search: Optional[str] = Query(None, alias="q"),
    limit: int = Query(25, le=100),
    db: AsyncSession = Depends(get_db),
):
    q = select(Member)
    if search and search.strip():
        term = f"%{search.strip()}%"
        q = q.where(or_(
            Member.name.ilike(term),
            Member.phone.ilike(term),
        ))
    q = q.order_by(Member.name).limit(limit)
    result = await db.execute(q)
    return result.scalars().all()


@router.get("/search", response_model=list[MemberOut])
async def search_members(
    q: str = Query(..., min_length=1),
    limit: int = Query(25, le=100),
    db: AsyncSession = Depends(get_db),
):
    term = f"%{q.strip()}%"
    query = select(Member).where(or_(
        Member.name.ilike(term),
        Member.phone.ilike(term),
    )).order_by(Member.name).limit(limit)
    result = await db.execute(query)
    return result.scalars().all()


@router.post("/import-excel")
async def import_members_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported.")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {str(e)}")

    sheet = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    inserted = 0
    updated = 0

    for r in range(2, sheet.max_row + 1):
        name = sheet.cell(r, 2).value
        phone = sheet.cell(r, 3).value
        address = sheet.cell(r, 4).value

        if name is not None and str(name).strip():
            name_clean = str(name).strip()
            phone_clean = str(phone).strip() if phone is not None else ""
            if phone_clean.endswith(".0"):
                phone_clean = phone_clean[:-2]
            address_clean = str(address).strip() if address is not None else ""

            res = await db.execute(select(Member).where(Member.name == name_clean))
            existing = res.scalars().first()

            if existing:
                if existing.phone != phone_clean or existing.address != address_clean:
                    existing.phone = phone_clean
                    existing.address = address_clean
                    updated += 1
            else:
                m = Member(
                    id=str(uuid.uuid4()),
                    name=name_clean,
                    phone=phone_clean,
                    address=address_clean,
                )
                db.add(m)
                inserted += 1

    await db.commit()
    return {
        "message": "Devotees imported successfully",
        "inserted": inserted,
        "updated": updated,
    }


@router.get("/{member_id}", response_model=MemberOut)
async def get_member(member_id: str, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Member).where(Member.id == member_id))
    member = result.scalar_one_or_none()
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    return member


@router.post("", response_model=MemberOut)
async def create_member(
    body: MemberCreate,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    member = Member(id=str(uuid.uuid4()), **body.model_dump())
    db.add(member)
    await db.commit()
    await db.refresh(member)
    return member


@router.put("/{member_id}", response_model=MemberOut)
async def update_member(
    member_id: str,
    body: MemberUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    result = await db.execute(select(Member).where(Member.id == member_id))
    member = result.scalar_one_or_none()
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    for k, v in body.model_dump(exclude_none=True).items():
        setattr(member, k, v)
    await db.commit()
    await db.refresh(member)
    return member


@router.delete("/{member_id}")
async def delete_member(
    member_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    result = await db.execute(select(Member).where(Member.id == member_id))
    member = result.scalar_one_or_none()
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    await db.delete(member)
    await db.commit()
    return {"message": "Devotee deleted successfully"}
